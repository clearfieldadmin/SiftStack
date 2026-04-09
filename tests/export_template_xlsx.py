"""Export a clean, simplified Excel template for demo/presentation.

Usage: python tests/export_template_xlsx.py <input_csv> [--output <path>]
"""

import argparse
import csv
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter


# (display_name, csv_field, format_type)
COLUMNS = [
    ("Property Address", "address", None),
    ("City", "city", None),
    ("State", "state", None),
    ("ZIP", "zip", None),
    ("Owner Name", "full_name", None),
    ("Date Added", "Date Added", None),
    ("Auction Date", "auction_date", None),
    ("Mailable", "mailable", None),
    ("USPS Verified", "dpv_match_code", None),
    ("Estimated Value", "estimated_value", "currency"),
    ("Equity %", "equity_percent", "pct"),
    ("Property Type", "property_type", None),
    ("Beds", "bedrooms", "int"),
    ("Baths", "bathrooms", "int"),
    ("Sq Ft", "sqft", "int"),
    ("Year Built", "year_built", "int"),
    ("MLS Status", "mls_status", None),
    ("Tax Delinquent", "tax_delinquent_amount", "currency"),
    ("Tax Years Owed", "tax_delinquent_years", "int"),
    ("Deceased", "owner_deceased", None),
]

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")


def main():
    parser = argparse.ArgumentParser(description="Export clean template Excel.")
    parser.add_argument("input_csv", help="Path to input CSV")
    parser.add_argument("--output", help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.input_csv, "r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    print(f"Loaded {len(rows)} records")

    wb = Workbook()
    ws = wb.active
    ws.title = "Knox Foreclosures"

    # Header row
    for c, (name, _, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for r, row in enumerate(rows, 2):
        for c, (_, field, fmt) in enumerate(COLUMNS, 1):
            val = row.get(field, "")
            cell = ws.cell(row=r, column=c)

            if fmt == "currency" and val:
                try:
                    cell.value = float(val.replace(",", "").replace("$", ""))
                    cell.number_format = "$#,##0"
                    continue
                except ValueError:
                    pass

            if fmt == "pct" and val:
                try:
                    v = float(val)
                    cell.value = v / 100 if v > 1 else v
                    cell.number_format = "0.0%"
                    continue
                except ValueError:
                    pass

            if fmt == "int" and val:
                try:
                    cell.value = int(float(val))
                    continue
                except ValueError:
                    pass

            cell.value = val

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto-fit column widths
    for col_idx in range(1, len(COLUMNS) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                max_len = max(max_len, min(len(str(cell.value or "")), 35))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Save
    if args.output:
        out_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y-%m-%d")
        out_path = os.path.join("output", f"knox_foreclosure_template_{timestamp}.xlsx")

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
