"""Export reenriched CSV to a multi-tab Excel workbook for quality review.

Usage: python tests/export_review_xlsx.py <input_csv> [--output <path>]
"""

import argparse
import csv
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter


# ── Column definitions per tab ────────────────────────────────────────

DECEASED_DM_COLS = [
    # Property
    "address", "city", "zip", "full_name", "tax_owner_name",
    "parcel_id", "mailable", "dpv_match_code",
    # Tax delinquency
    "tax_delinquent_amount", "tax_delinquent_years",
    # Property value / equity
    "estimated_value", "estimated_equity", "equity_percent",
    "property_type", "bedrooms", "bathrooms", "sqft", "year_built", "lot_size",
    # MLS
    "mls_status", "mls_listing_price", "mls_last_sold_date", "mls_last_sold_price",
    # Deceased / DM
    "date_of_death", "obituary_url", "obituary_source_type",
    "decision_maker_name", "decision_maker_relationship",
    "decision_maker_status", "decision_maker_source",
    "decision_maker_street", "decision_maker_city",
    "decision_maker_state", "decision_maker_zip",
    "dm_confidence", "dm_confidence_reason",
    "heirs_verified_living", "heirs_verified_deceased", "heirs_unverified",
    "decision_maker_2_name", "decision_maker_2_relationship",
    "decision_maker_3_name", "decision_maker_3_relationship",
    "missing_data_flags", "source_url",
]

LIVING_COLS = [
    "address", "city", "zip", "full_name", "tax_owner_name",
    "parcel_id", "tax_delinquent_amount", "tax_delinquent_years",
    "estimated_value", "estimated_equity", "equity_percent",
    "dpv_match_code", "mailable", "deceased_indicator",
    "property_type", "bedrooms", "bathrooms", "sqft", "year_built",
    "lot_size", "source_url",
]

MISSING_COLS = [
    "address", "city", "zip", "full_name", "tax_owner_name",
    "parcel_id", "mailable", "dpv_match_code",
    "owner_deceased", "decision_maker_name", "decision_maker_street",
    "missing_data_flags", "dm_confidence", "source_url",
]

# ── Styles ────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FONT = Font(color="CC0000", bold=True)
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
CURRENCY_FMT = '#,##0.00'
PCT_FMT = '0.0%'


def _read_csv(path: str) -> list[dict]:
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _style_header(ws):
    """Bold blue header, frozen, auto-filter."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_width(ws, max_width=40):
    """Auto-fit column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)


def _write_data_sheet(ws, rows: list[dict], columns: list[str]):
    """Write column subset to a worksheet with formatting."""
    # Header
    for c, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=c, value=col_name)

    # Data
    currency_cols = {"tax_delinquent_amount", "estimated_value", "estimated_equity",
                     "mls_listing_price", "mls_last_sold_price"}

    for r, row in enumerate(rows, 2):
        for c, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            cell = ws.cell(row=r, column=c)

            # Numeric conversion for currency/number columns
            if col_name in currency_cols and val:
                try:
                    cell.value = float(val.replace(",", "").replace("$", ""))
                    cell.number_format = CURRENCY_FMT
                    continue
                except ValueError:
                    pass

            if col_name == "equity_percent" and val:
                try:
                    cell.value = float(val) / 100 if float(val) > 1 else float(val)
                    cell.number_format = PCT_FMT
                    continue
                except ValueError:
                    pass

            if col_name in ("tax_delinquent_years", "bedrooms", "bathrooms",
                            "sqft", "year_built", "heirs_verified_living",
                            "heirs_verified_deceased", "heirs_unverified") and val:
                try:
                    cell.value = int(float(val))
                    continue
                except ValueError:
                    pass

            # Hyperlinks for URL columns
            if col_name in ("obituary_url", "source_url") and val and val.startswith("http"):
                cell.value = val
                cell.font = LINK_FONT
                cell.hyperlink = val
                continue

            cell.value = val

    _style_header(ws)
    _auto_width(ws)


def _apply_dm_formatting(ws, columns: list[str]):
    """Conditional formatting for DM quality columns."""
    conf_col = None
    source_col = None
    for c, col_name in enumerate(columns, 1):
        if col_name == "dm_confidence":
            conf_col = c
        if col_name == "decision_maker_source":
            source_col = c

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if conf_col:
            cell = row[conf_col - 1]
            val = str(cell.value or "").lower()
            if val in ("high", "medium"):
                cell.fill = GREEN_FILL
            elif val == "low":
                cell.fill = YELLOW_FILL

        if source_col:
            cell = row[source_col - 1]
            val = str(cell.value or "")
            if val == "estate_fallback":
                cell.font = RED_FONT


def _write_summary(ws, rows: list[dict]):
    """Write summary dashboard tab."""
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30

    def _add(r, label, value, note=""):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        if note:
            ws.cell(row=r, column=3, value=note).font = Font(color="666666", italic=True)

    total = len(rows)
    deceased = [r for r in rows if r.get("owner_deceased") == "yes"]
    with_dm = [r for r in deceased if r.get("decision_maker_name")]
    with_dm_addr = [r for r in deceased if r.get("decision_maker_street")]
    mailable = [r for r in rows if r.get("mailable") == "yes"]
    estate = [r for r in deceased if (r.get("decision_maker_name") or "").startswith("Estate of")]

    # Tax stats
    tax_amounts = []
    for r in rows:
        try:
            amt = float((r.get("tax_delinquent_amount") or "0").replace(",", "").replace("$", ""))
            if amt > 0:
                tax_amounts.append(amt)
        except ValueError:
            pass

    dpv_y = sum(1 for r in rows if r.get("dpv_match_code") == "Y")

    # DM source breakdown
    dm_sources = {}
    for r in deceased:
        s = r.get("decision_maker_source", "none")
        dm_sources[s] = dm_sources.get(s, 0) + 1

    # DM confidence breakdown
    dm_conf = {}
    for r in deceased:
        c = r.get("dm_confidence", "none")
        dm_conf[c] = dm_conf.get(c, 0) + 1

    r = 1
    ws.cell(row=r, column=1, value="Quality Review Summary").font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(color="666666")
    r += 2

    # Section: Overview
    ws.cell(row=r, column=1, value="OVERVIEW").font = Font(bold=True, size=12, color="1F4E79")
    r += 1
    _add(r, "Total Records", total); r += 1
    _add(r, "Mailable", f"{len(mailable)}/{total} ({len(mailable)*100//total}%)"); r += 1
    _add(r, "Smarty DPV-Y (USPS confirmed)", f"{dpv_y}/{total} ({dpv_y*100//total}%)"); r += 1
    r += 1

    # Section: Deceased + DM
    ws.cell(row=r, column=1, value="DECEASED OWNER COVERAGE").font = Font(bold=True, size=12, color="1F4E79")
    r += 1
    _add(r, "Deceased Confirmed", f"{len(deceased)}/{total} ({len(deceased)*100//total}%)"); r += 1
    _add(r, "Decision-Maker Identified", f"{len(with_dm)}/{len(deceased)} ({len(with_dm)*100//max(len(deceased),1)}%)",
         "Target: 100%"); r += 1
    _add(r, "DM with Address", f"{len(with_dm_addr)}/{len(deceased)} ({len(with_dm_addr)*100//max(len(deceased),1)}%)",
         "Target: 100%"); r += 1
    _add(r, "Estate-of Fallback", len(estate), "No living relative found, mail to property"); r += 1
    r += 1

    # Section: DM Sources
    ws.cell(row=r, column=1, value="DM ADDRESS SOURCES").font = Font(bold=True, size=12, color="1F4E79")
    r += 1
    for source, count in sorted(dm_sources.items(), key=lambda x: -x[1]):
        _add(r, f"  {source}", count); r += 1
    r += 1

    # Section: DM Confidence
    ws.cell(row=r, column=1, value="DM CONFIDENCE").font = Font(bold=True, size=12, color="1F4E79")
    r += 1
    for conf, count in sorted(dm_conf.items(), key=lambda x: -x[1]):
        _add(r, f"  {conf}", count); r += 1
    r += 1

    # Section: Tax Delinquency
    ws.cell(row=r, column=1, value="TAX DELINQUENCY").font = Font(bold=True, size=12, color="1F4E79")
    r += 1
    _add(r, "Tax-Delinquent Records", len(tax_amounts)); r += 1
    if tax_amounts:
        avg_amt = sum(tax_amounts) / len(tax_amounts)
        _add(r, "Average Delinquent Amount", f"${avg_amt:,.2f}"); r += 1
        _add(r, "Total Delinquent Amount", f"${sum(tax_amounts):,.2f}"); r += 1
        _add(r, "Max Delinquent Amount", f"${max(tax_amounts):,.2f}"); r += 1

    # Freeze and style
    ws.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser(description="Export quality review Excel workbook.")
    parser.add_argument("input_csv", help="Path to reenriched CSV")
    parser.add_argument("--output", help="Output .xlsx path (default: auto-generated in output/)")
    args = parser.parse_args()

    rows = _read_csv(args.input_csv)
    print(f"Loaded {len(rows)} records from {args.input_csv}")

    wb = Workbook()

    # ── Tab 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, rows)

    # ── Tab 2: Deceased + DM ──
    deceased = [r for r in rows if r.get("owner_deceased") == "yes"]
    ws_deceased = wb.create_sheet("Deceased + DM")
    _write_data_sheet(ws_deceased, deceased, DECEASED_DM_COLS)
    _apply_dm_formatting(ws_deceased, DECEASED_DM_COLS)
    print(f"  Deceased + DM: {len(deceased)} rows")

    # ── Tab 3: Estate Fallbacks ──
    estate = [r for r in deceased if (r.get("decision_maker_name") or "").startswith("Estate of")]
    ws_estate = wb.create_sheet("Estate Fallbacks")
    _write_data_sheet(ws_estate, estate, DECEASED_DM_COLS)
    _apply_dm_formatting(ws_estate, DECEASED_DM_COLS)
    print(f"  Estate Fallbacks: {len(estate)} rows")

    # ── Tab 4: Living / Not Deceased ──
    living = [r for r in rows if r.get("owner_deceased") != "yes"]
    ws_living = wb.create_sheet("Living - Not Deceased")
    _write_data_sheet(ws_living, living, LIVING_COLS)
    print(f"  Living / Not Deceased: {len(living)} rows")

    # ── Tab 5: Missing Data ──
    missing = [
        r for r in rows
        if not r.get("address")
        or not r.get("city")
        or not r.get("zip")
        or not r.get("full_name")
        or r.get("mailable") != "yes"
        or r.get("missing_data_flags")
        or (r.get("owner_deceased") == "yes" and not r.get("decision_maker_street"))
    ]
    ws_missing = wb.create_sheet("Missing Data")
    _write_data_sheet(ws_missing, missing, MISSING_COLS)
    print(f"  Missing Data: {len(missing)} rows")

    # ── Save ──
    if args.output:
        out_path = args.output
    else:
        timestamp = datetime.now().strftime("%Y-%m-%d")
        out_path = os.path.join("output", f"quality_review_{timestamp}.xlsx")

    wb.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
