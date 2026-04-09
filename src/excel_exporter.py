"""Export enrichment CSV to a multi-sheet Excel review workbook.

Sheets:
  1. Dashboard  — summary stats at a glance
  2. Deceased Review — focused view of deceased records with DM info
  3. All Records — full dataset with curated columns

Usage:
  python src/excel_exporter.py [csv_path]
"""

import csv
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ─────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
STAT_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
STAT_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
STAT_PCT_FONT = Font(name="Calibri", bold=True, size=13, color="2F5496")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="Calibri", bold=True, color="006100")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(name="Calibri", bold=True, color="9C6500")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Calibri", bold=True, color="9C0006")
DARK_RED_FILL = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
DARK_RED_FONT = Font(name="Calibri", bold=True, color="632523")
GRAY_FONT = Font(name="Calibri", color="808080")
DECEASED_ROW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)

# ── Column definitions ─────────────────────────────────────────────

# Deceased Review sheet columns: (display_name, csv_column)
DECEASED_COLUMNS = [
    ("Owner Name", "full_name"),
    ("Address", "address"),
    ("City", "city"),
    ("Date of Death", "date_of_death"),
    ("Decision Maker", "decision_maker_name"),
    ("DM Relationship", "decision_maker_relationship"),
    ("DM Status", "decision_maker_status"),
    ("DM Confidence", "dm_confidence"),
    ("Confidence Reason", "dm_confidence_reason"),
    ("DM Street", "decision_maker_street"),
    ("DM City", "decision_maker_city"),
    ("DM State", "decision_maker_state"),
    ("DM ZIP", "decision_maker_zip"),
    ("2nd DM", "decision_maker_2_name"),
    ("2nd DM Relationship", "decision_maker_2_relationship"),
    ("3rd DM", "decision_maker_3_name"),
    ("3rd DM Relationship", "decision_maker_3_relationship"),
    ("Source Type", "obituary_source_type"),
    ("Obituary URL", "obituary_url"),
    ("Data Flags", "missing_data_flags"),
    ("Tax Owner Name", "tax_owner_name"),
    ("Est. Value", "estimated_value"),
    ("Est. Equity", "estimated_equity"),
    ("Heir Map", "_heir_map"),  # special: populated by _build_heir_map_note()
]

# All Records sheet columns: (display_name, csv_column)
ALL_COLUMNS = [
    ("Owner Name", "full_name"),
    ("Address", "address"),
    ("City", "city"),
    ("State", "state"),
    ("ZIP", "zip"),
    ("Notice Type", "notice_type"),
    ("County", "county"),
    ("Parcel ID", "parcel_id"),
    ("Est. Value", "estimated_value"),
    ("Est. Equity", "estimated_equity"),
    ("Equity %", "equity_percent"),
    ("Property Type", "property_type"),
    ("Tax Delinquent $", "tax_delinquent_amount"),
    ("Tax Delinq. Years", "tax_delinquent_years"),
    ("Deceased?", "owner_deceased"),
    ("Date of Death", "date_of_death"),
    ("Decision Maker", "decision_maker_name"),
    ("DM Relationship", "decision_maker_relationship"),
    ("DM Status", "decision_maker_status"),
    ("DM Confidence", "dm_confidence"),
    ("DM Street", "decision_maker_street"),
    ("DM City", "decision_maker_city"),
    ("DM State", "decision_maker_state"),
    ("DM ZIP", "decision_maker_zip"),
    ("Obituary URL", "obituary_url"),
    ("Mailable", "mailable"),
    ("Heir Map", "_heir_map"),  # special: populated by _build_heir_map_note()
]


def _load_csv(path: str) -> list[dict]:
    """Load CSV into list of row dicts."""
    with open(path, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _auto_width(ws, max_width=40):
    """Set column widths based on content, capped at max_width."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        width = min(max_len + 3, max_width)
        ws.column_dimensions[col_letter].width = max(width, 10)


def _style_header_row(ws, num_cols):
    """Apply header styling to row 1."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _build_heir_map_note(row: dict) -> tuple[str, str]:
    """Build heir map summary + detailed popup note from CSV row data.

    Returns (summary_text, full_note_text) where:
      - summary_text goes in the cell (visible in column)
      - full_note_text goes in a Comment popup (hover to see)
    """
    dm1_name = row.get("decision_maker_name", "")
    dm1_rel = row.get("decision_maker_relationship", "")
    dm1_status = row.get("decision_maker_status", "")
    dm1_source = row.get("decision_maker_source", "")
    dm1_street = row.get("decision_maker_street", "")
    dm1_city = row.get("decision_maker_city", "")
    dm1_state = row.get("decision_maker_state", "")
    dm1_zip = row.get("decision_maker_zip", "")

    dm2_name = row.get("decision_maker_2_name", "")
    dm2_rel = row.get("decision_maker_2_relationship", "")
    dm2_status = row.get("decision_maker_2_status", "")

    dm3_name = row.get("decision_maker_3_name", "")
    dm3_rel = row.get("decision_maker_3_relationship", "")
    dm3_status = row.get("decision_maker_3_status", "")

    # Status icons
    def _icon(status):
        if status == "verified_living":
            return "VERIFIED LIVING"
        elif status == "deceased":
            return "DECEASED"
        elif status == "unverified":
            return "unverified"
        return ""

    def _short_icon(status):
        if status == "verified_living":
            return "verified"
        elif status == "deceased":
            return "X"
        elif status == "unverified":
            return "?"
        return ""

    # Count total heirs shown
    heir_count = sum(1 for n in [dm1_name, dm2_name, dm3_name] if n)

    # ── Summary (cell value) ──
    if dm1_name:
        extra = f" + {heir_count - 1} more" if heir_count > 1 else ""
        summary = f"{dm1_name} ({dm1_rel}) [{_short_icon(dm1_status)}]{extra}"
    elif "no_dm_possible" in row.get("missing_data_flags", ""):
        summary = "No heirs found"
    else:
        summary = "No DM identified"

    # ── Full note (popup) ──
    lines = []

    # Decedent section
    owner = row.get("full_name", "")
    dod = row.get("date_of_death", "")
    source = row.get("obituary_source_type", "")
    lines.append("DECEDENT")
    lines.append(f"  Name:    {owner}")
    if dod:
        lines.append(f"  Died:    {dod}")
    lines.append(f"  Source:  {source}")
    lines.append("")

    # Heir map section
    lines.append("HEIR MAP")
    if dm1_name:
        lines.append(f"  * 1. {dm1_name} ({dm1_rel}) -- {_icon(dm1_status)}")
        if dm1_source and dm1_source != "obituary_survivors":
            lines.append(f"       Source: {dm1_source}")
        if dm1_street:
            addr_parts = [dm1_street]
            if dm1_city:
                addr_parts.append(dm1_city)
            if dm1_state:
                addr_parts.append(dm1_state)
            if dm1_zip:
                addr_parts.append(dm1_zip)
            lines.append(f"       Address: {', '.join(addr_parts)}")
    if dm2_name:
        lines.append(f"    2. {dm2_name} ({dm2_rel}) -- {_icon(dm2_status)}")
    if dm3_name:
        lines.append(f"    3. {dm3_name} ({dm3_rel}) -- {_icon(dm3_status)}")
    if not dm1_name:
        flags = row.get("missing_data_flags", "")
        if "no_dm_possible" in flags:
            lines.append("  (no family members identifiable)")
        elif "no_survivors" in flags:
            lines.append("  (no survivors in obituary)")
        else:
            lines.append("  (no decision-maker identified)")
    lines.append("")

    # Verification section
    depth = row.get("heir_search_depth", "0")
    living = row.get("heirs_verified_living", "0")
    deceased_count = row.get("heirs_verified_deceased", "0")
    unverified = row.get("heirs_unverified", "0")
    conf = row.get("dm_confidence", "")
    reason = row.get("dm_confidence_reason", "")

    lines.append("VERIFICATION")
    lines.append(f"  Depth: {depth} | Living: {living} | Deceased: {deceased_count} | Unverified: {unverified}")
    if conf:
        lines.append(f"  Confidence: {conf.upper()}")
    if reason:
        lines.append(f"  Reason: {reason}")

    # Flags section
    flags = row.get("missing_data_flags", "")
    if flags:
        lines.append("")
        lines.append(f"FLAGS: {flags}")

    note_text = "\n".join(lines)
    return summary, note_text


def _build_dashboard(wb, rows):
    """Build the Dashboard summary sheet."""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "2F5496"

    total = len(rows)
    deceased_rows = [r for r in rows if r.get("owner_deceased") == "yes"]
    deceased = len(deceased_rows)
    mailable = sum(1 for r in rows if r.get("mailable") == "yes")
    full_page = sum(1 for r in deceased_rows if r.get("obituary_source_type") == "full_page")
    snippet = sum(1 for r in deceased_rows if r.get("obituary_source_type") == "snippet")
    with_dm = sum(1 for r in deceased_rows if r.get("decision_maker_name"))
    dm_verified = sum(1 for r in deceased_rows if r.get("decision_maker_status") == "verified_living")
    dm_from_tax = sum(1 for r in deceased_rows if r.get("decision_maker_source") == "tax_record_joint_owner")
    dm_snippet = sum(1 for r in deceased_rows if "dm_from_targeted_snippet" in r.get("missing_data_flags", ""))
    no_dm = sum(1 for r in deceased_rows if "no_dm_possible" in r.get("missing_data_flags", ""))
    with_dod = sum(1 for r in deceased_rows if r.get("date_of_death"))
    with_url = sum(1 for r in deceased_rows if r.get("obituary_url"))
    with_dm2 = sum(1 for r in deceased_rows if r.get("decision_maker_2_name"))
    with_dm3 = sum(1 for r in deceased_rows if r.get("decision_maker_3_name"))
    high_conf = sum(1 for r in deceased_rows if r.get("dm_confidence") == "high")
    med_conf = sum(1 for r in deceased_rows if r.get("dm_confidence") == "medium")
    low_conf = sum(1 for r in deceased_rows if r.get("dm_confidence") == "low")
    none_conf = sum(1 for r in deceased_rows if r.get("dm_confidence") == "none")
    dm_with_addr = sum(1 for r in deceased_rows if r.get("decision_maker_street"))
    sift_ready = sum(1 for r in deceased_rows if r.get("decision_maker_name") and r.get("decision_maker_street"))

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Obituary Enrichment Review"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="888888")

    def _stat(row, label, value, pct=None):
        ws.cell(row=row, column=1, value=label).font = STAT_LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = STAT_VALUE_FONT
        if pct is not None:
            ws.cell(row=row, column=3, value=f"({pct:.0f}%)").font = STAT_PCT_FONT

    # Section: Dataset Overview
    r = 4
    ws.cell(row=r, column=1, value="DATASET OVERVIEW").font = SUBTITLE_FONT
    r += 1
    _stat(r, "Total Records", total); r += 1
    _stat(r, "Mailable", mailable, 100 * mailable / total if total else 0); r += 1
    _stat(r, "Confirmed Deceased", deceased, 100 * deceased / total if total else 0); r += 1

    # Section: Obituary Match Quality
    r += 1
    ws.cell(row=r, column=1, value="OBITUARY MATCH QUALITY").font = SUBTITLE_FONT
    r += 1
    _stat(r, "Full-Page Matches", full_page); r += 1
    _stat(r, "Snippet Matches", snippet); r += 1
    _stat(r, "Has Date of Death", with_dod, 100 * with_dod / deceased if deceased else 0); r += 1
    _stat(r, "Has Obituary URL", with_url, 100 * with_url / deceased if deceased else 0); r += 1

    # Section: Decision-Maker Coverage
    r += 1
    ws.cell(row=r, column=1, value="DECISION-MAKER COVERAGE").font = SUBTITLE_FONT
    r += 1
    _stat(r, "DM Identified", with_dm, 100 * with_dm / deceased if deceased else 0); r += 1
    _stat(r, "DM Verified Living", dm_verified); r += 1
    _stat(r, "DM from Tax Record (Joint Owner)", dm_from_tax); r += 1
    _stat(r, "DM from Targeted Snippet Search", dm_snippet); r += 1
    _stat(r, "No DM Possible (Flagged)", no_dm); r += 1
    _stat(r, "Has 2nd DM", with_dm2); r += 1
    _stat(r, "Has 3rd DM", with_dm3); r += 1

    # Section: Confidence Breakdown
    r += 1
    ws.cell(row=r, column=1, value="DM CONFIDENCE BREAKDOWN").font = SUBTITLE_FONT
    r += 1
    _stat(r, "High Confidence", high_conf)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=13, color="006100")
    r += 1
    _stat(r, "Medium Confidence", med_conf)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=13, color="9C6500")
    r += 1
    _stat(r, "Low Confidence", low_conf)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=13, color="9C0006")
    r += 1
    _stat(r, "None (No DM Possible)", none_conf)
    ws.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=13, color="632523")

    # Section: DM Address / Sift Readiness
    r += 2
    ws.cell(row=r, column=1, value="DM ADDRESS & SIFT READINESS").font = SUBTITLE_FONT
    r += 1
    _stat(r, "DM Has Mailing Address", dm_with_addr, 100 * dm_with_addr / with_dm if with_dm else 0); r += 1
    _stat(r, "Sift-Ready (DM + Address)", sift_ready, 100 * sift_ready / deceased if deceased else 0); r += 1

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10


def _build_deceased_review(wb, rows):
    """Build the Deceased Review sheet with focused columns and formatting."""
    ws = wb.create_sheet("Deceased Review")
    ws.sheet_properties.tabColor = "C00000"

    deceased_rows = [r for r in rows if r.get("owner_deceased") == "yes"]

    # Sort: none → low → medium → high (problem records first)
    conf_order = {"none": 0, "low": 1, "medium": 2, "high": 3, "": 4}
    deceased_rows.sort(key=lambda r: conf_order.get(r.get("dm_confidence", ""), 4))

    # Headers
    for col_idx, (display_name, _) in enumerate(DECEASED_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=display_name)
    _style_header_row(ws, len(DECEASED_COLUMNS))

    # Find Heir Map column index
    heir_map_col = next(
        (i for i, (_, k) in enumerate(DECEASED_COLUMNS, 1) if k == "_heir_map"),
        None,
    )

    # Data rows
    for row_idx, row in enumerate(deceased_rows, 2):
        for col_idx, (_, csv_col) in enumerate(DECEASED_COLUMNS, 1):
            if csv_col == "_heir_map":
                continue  # populated separately below
            value = row.get(csv_col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

        # Heir Map column: summary + popup comment
        if heir_map_col:
            summary, note_text = _build_heir_map_note(row)
            hm_cell = ws.cell(row=row_idx, column=heir_map_col, value=summary)
            hm_cell.border = THIN_BORDER
            hm_cell.alignment = Alignment(vertical="center")
            comment = Comment(note_text, "Heir Map")
            comment.width = 400
            comment.height = 350
            hm_cell.comment = comment

        # Obituary URL as hyperlink (column 19)
        url_val = row.get("obituary_url", "")
        if url_val:
            url_cell = ws.cell(row=row_idx, column=19)
            try:
                url_cell.hyperlink = url_val
                url_cell.font = Font(name="Calibri", color="0563C1", underline="single")
                url_cell.value = "View Obituary"
            except Exception:
                pass

        # DM Status coloring (column 7)
        dm_status = row.get("decision_maker_status", "")
        status_cell = ws.cell(row=row_idx, column=7)
        if dm_status == "verified_living":
            status_cell.fill = GREEN_FILL
            status_cell.font = GREEN_FONT
        elif dm_status == "deceased":
            status_cell.fill = RED_FILL
            status_cell.font = RED_FONT
        elif dm_status == "unverified":
            status_cell.font = GRAY_FONT

        # DM Name coloring (column 5)
        dm_name = row.get("decision_maker_name", "")
        name_cell = ws.cell(row=row_idx, column=5)
        if not dm_name:
            name_cell.value = "(none)"
            name_cell.font = Font(name="Calibri", italic=True, color="CC0000")
        elif dm_status == "verified_living":
            name_cell.font = GREEN_FONT
        elif dm_status == "deceased":
            name_cell.font = RED_FONT

        # Confidence coloring (column 8)
        conf = row.get("dm_confidence", "")
        conf_cell = ws.cell(row=row_idx, column=8)
        if conf == "high":
            conf_cell.fill = GREEN_FILL
            conf_cell.font = GREEN_FONT
        elif conf == "medium":
            conf_cell.fill = YELLOW_FILL
            conf_cell.font = YELLOW_FONT
        elif conf == "low":
            conf_cell.fill = RED_FILL
            conf_cell.font = RED_FONT
        elif conf == "none":
            conf_cell.fill = DARK_RED_FILL
            conf_cell.font = DARK_RED_FONT

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DECEASED_COLUMNS))}1"
    _auto_width(ws)
    # Widen Heir Map column for summary text
    if heir_map_col:
        ws.column_dimensions[get_column_letter(heir_map_col)].width = 45


def _build_all_records(wb, rows):
    """Build the All Records sheet with curated columns."""
    ws = wb.create_sheet("All Records")
    ws.sheet_properties.tabColor = "4472C4"

    # Headers
    for col_idx, (display_name, _) in enumerate(ALL_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=display_name)
    _style_header_row(ws, len(ALL_COLUMNS))

    # Find Heir Map column index
    all_heir_map_col = next(
        (i for i, (_, k) in enumerate(ALL_COLUMNS, 1) if k == "_heir_map"),
        None,
    )

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        is_deceased = row.get("owner_deceased") == "yes"
        for col_idx, (_, csv_col) in enumerate(ALL_COLUMNS, 1):
            if csv_col == "_heir_map":
                continue  # populated separately below
            value = row.get(csv_col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            if is_deceased:
                cell.fill = DECEASED_ROW_FILL

        # Heir Map column (deceased rows only)
        if all_heir_map_col and is_deceased:
            summary, note_text = _build_heir_map_note(row)
            hm_cell = ws.cell(row=row_idx, column=all_heir_map_col, value=summary)
            hm_cell.border = THIN_BORDER
            hm_cell.alignment = Alignment(vertical="center")
            hm_cell.fill = DECEASED_ROW_FILL
            comment = Comment(note_text, "Heir Map")
            comment.width = 400
            comment.height = 350
            hm_cell.comment = comment

        # Obituary URL hyperlink (column 25)
        url_val = row.get("obituary_url", "")
        if url_val:
            url_cell = ws.cell(row=row_idx, column=25)
            try:
                url_cell.hyperlink = url_val
                url_cell.font = Font(name="Calibri", color="0563C1", underline="single")
                url_cell.value = "View"
            except Exception:
                pass

        # Deceased column coloring (column 15)
        if is_deceased:
            ws.cell(row=row_idx, column=15).font = RED_FONT

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_COLUMNS))}1"
    _auto_width(ws)
    # Widen Heir Map column
    if all_heir_map_col:
        ws.column_dimensions[get_column_letter(all_heir_map_col)].width = 45


# Sift Upload sheet: ready-to-import format for deceased records
SIFT_UPLOAD_COLUMNS = [
    ("full_name", "decision_maker_name"),       # DM name becomes the contact
    ("address", "address"),                      # Property address (unchanged)
    ("city", "city"),
    ("state", "state"),
    ("zip", "zip"),
    ("Owner Street", "decision_maker_street"),   # DM's mailing address
    ("Owner City", "decision_maker_city"),
    ("Owner State", "decision_maker_state"),
    ("Owner ZIP Code", "decision_maker_zip"),
    ("Date Added", "Date Added"),
    ("notice_type", "notice_type"),
    ("county", "county"),
    ("parcel_id", "parcel_id"),
    ("estimated_value", "estimated_value"),
    ("estimated_equity", "estimated_equity"),
]


def _build_sift_upload(wb, rows):
    """Build the Sift Upload sheet — deceased records with DM name/address swapped.

    Only includes rows where both DM name and DM street are populated.
    This sheet can be exported directly to Sift CSV for marketing flows.
    """
    ws = wb.create_sheet("Sift Upload")
    ws.sheet_properties.tabColor = "00B050"

    # Filter: deceased + DM name + DM street all present
    upload_rows = [
        r for r in rows
        if r.get("owner_deceased") == "yes"
        and r.get("decision_maker_name")
        and r.get("decision_maker_street")
    ]

    # Headers
    for col_idx, (display_name, _) in enumerate(SIFT_UPLOAD_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=display_name)
    _style_header_row(ws, len(SIFT_UPLOAD_COLUMNS))

    # Data rows
    for row_idx, row in enumerate(upload_rows, 2):
        for col_idx, (_, csv_col) in enumerate(SIFT_UPLOAD_COLUMNS, 1):
            value = row.get(csv_col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    if upload_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(SIFT_UPLOAD_COLUMNS))}{len(upload_rows) + 1}"
    _auto_width(ws)


def export_review_workbook(csv_path: str, output_path: str = None) -> str:
    """Read enrichment CSV and produce a multi-sheet Excel review workbook."""
    rows = _load_csv(csv_path)
    if not rows:
        print(f"ERROR: No data in {csv_path}")
        sys.exit(1)

    wb = Workbook()

    _build_dashboard(wb, rows)
    _build_deceased_review(wb, rows)
    _build_all_records(wb, rows)
    _build_sift_upload(wb, rows)

    if not output_path:
        base_dir = os.path.dirname(csv_path)
        timestamp = datetime.now().strftime("%Y-%m-%d")
        output_path = os.path.join(base_dir, f"knox_tax_sale_review_{timestamp}.xlsx")

    wb.save(output_path)
    print(f"Excel workbook saved: {output_path}")
    return output_path


if __name__ == "__main__":
    default_csv = "output/knox_tax_sale_obituary_2026-03-03_124916.csv"
    csv_path = sys.argv[1] if len(sys.argv) > 1 else default_csv
    export_review_workbook(csv_path)
