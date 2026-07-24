"""export_excel.py - render call coaching reports into one clean Excel workbook.

Parses every per-call report .md in a reports directory. The primary data source
is the report's required SCORES JSON footer (rubric v1.2+); markdown tables are
the fallback for older reports. Full calls and short calls are DIFFERENT report
types on different scales, so they land on separate sheets and are never mixed
in one ranking column.

Sheets:
  Summary          full calls only: linked call ID, caller, category scores,
                   total /100, band (color coded)
  Short Calls      short calls only: opener /5, conversion /5, attempt made,
                   the one fix
  Coaching Detail  one row per coaching item (pillars, strengths, improvements,
                   what happened, fix, drill) for every call
  Caller Scorecard per-caller: full-call stats + short-call stats side by side
  Criterion Scores per-criterion grid (emitted only when most reports carry
                   criterion data)
  Old vs New       optional (--compare <old_dir>), full calls only

USAGE (from SiftStack root, venv python):
  python src/call_coaching/export_excel.py                              # cold_call reports
  python src/call_coaching/export_excel.py --dir output/call_coaching/reports/cold_call_v2 \
      --caller tinaa --out output/call_coaching/reports/tinaa.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DIR = ROOT / "output" / "call_coaching" / "reports" / "cold_call"

NAVY = "1F3864"
BAND_FILLS = {
    "Elite": "70AD47",
    "Strong": "C6E0B4",
    "Developing": "FFE699",
    "Needs Work": "F4B183",
    "Retrain": "FF7C80",
    "FAIL": "FF0000",
}

SECTION_RE = re.compile(r"^[A-Z][A-Z0-9 /()3']+$")


def _grab(pattern: str, text: str) -> str:
    m = re.search(pattern, text, re.M)
    return m.group(1).strip() if m else ""


def _numbered_items(block: str) -> list[str]:
    items: list[str] = []
    for line in block.splitlines():
        s = line.strip()
        if re.match(r"^\d\.", s):
            items.append(re.sub(r"^\d\.\s*", "", s))
        elif items and s and not SECTION_RE.match(s):
            items[-1] += " " + s
    return [re.sub(r"\*+", "", i).strip() for i in items[:3]]


def _section(text: str, header_prefix: str) -> str:
    lines, out, active = text.splitlines(), [], False
    for line in lines:
        s = line.strip()
        if not active and s.startswith(header_prefix):
            active = True
            inline = s[len(header_prefix):].lstrip(": ").strip()
            if inline:
                out.append(inline)
            continue
        if active:
            if (SECTION_RE.match(s) and len(s) > 8) or s.startswith("```"):
                break
            out.append(line)
    return "\n".join(out).strip()


def _scores_json(text: str) -> dict:
    """Last parseable JSON object containing call_id (the required footer)."""
    candidates = re.findall(r"```json\s*(\{.*?\})\s*```", text, re.S)
    candidates += re.findall(r"^(\{\"call_id\".*?\})\s*$", text, re.M | re.S)
    for raw in reversed(candidates):
        try:
            obj = json.loads(raw)
            if isinstance(obj, dict) and "call_id" in obj:
                return obj
        except (ValueError, TypeError):
            continue
    return {}


def parse_report(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")
    m = re.match(r"^(\d+)_(.+)$", path.stem)
    call_id = m.group(1) if m else path.stem
    caller = (m.group(2) if m else "").replace("_", " ").title()
    js = _scores_json(text)

    link = js.get("recording_url") or ""
    if not link:
        lm = re.search(r"\[\d{6,}\]\((https?://[^)]+)\)", text)
        link = lm.group(1) if lm else ""

    # markdown tables: category names + scores, criterion rows (fallback + names)
    cats_md: dict[str, str] = {}
    crit_names: dict[str, str] = {}
    crits: dict[str, object] = {}
    for line in text.splitlines():
        m2 = re.match(r"^\|\s*(\d(?:\.\d)?)\.?\s+([A-Za-z\"'][^|]*?)\s*\|(.+)\|\s*$", line)
        if not m2:
            continue
        num, name = m2.group(1), re.sub(r"\s*\([^)]*\)\s*$", "", m2.group(2))
        score = ""
        for cell in (c.strip() for c in m2.group(3).split("|")):
            if re.fullmatch(r"[0-5](\.\d+)?", cell) or cell.upper().startswith("N/A"):
                score = cell
                break
        if "." in num:
            crit_names[num] = name
            crits[num] = score
        else:
            cats_md[f"{num}. {name}"] = score
    for k, v in (js.get("criteria") or {}).items():
        crits[k] = "N/A" if v is None else v

    total = js.get("total", "")
    if total in ("", None) and js.get("call_type") != "short":
        tm = re.search(r"^\|\s*TOTAL[^|]*\|[^\n]*?([\d.]+)\s*/\s*100", text, re.M)
        total = tm.group(1) if tm else ""
    band = js.get("band") or _grab(r"Grade [Bb]and:\s*([^\n]+)", text)
    band = re.sub(r"\s*\([^)]*\)\s*$", "", (band or "").replace("*", "").strip())

    call_type = js.get("call_type") or ("short" if "SHORT CALL REPORT" in text else "full")
    dur = js.get("duration_seconds")
    length = f"{dur // 60}:{dur % 60:02d}" if isinstance(dur, int) else \
        _grab(r"Call length:\s*([^\n]+?)(?:\s{2,}|$)", text)

    fix = _section(text, "ONE FIX")
    drill = _section(text, "ONE DRILL")
    return {
        "call_id": str(js.get("call_id") or call_id),
        "caller": caller or str(js.get("caller") or "").title(),
        "call_type": call_type,
        "link": link,
        "datetime": _grab(r"Date/Time:\s*([^\n]+?)(?:\s{2,}Lead|$)", text),
        "length": length,
        "outcome": str(js.get("outcome") or _grab(r"Outcome:\s*([^\n]+)", text)),
        "autofail": str(js.get("auto_fail") or _grab(r"Auto-fail check:\s*([^\n]+)", text)).split(" ")[0],
        "cats_md": cats_md,
        "cats_js": js.get("categories") or {},
        "crits": crits,
        "crit_names": crit_names,
        "total": total,
        "band": band if call_type == "full" else "",
        "opener": js.get("opener", ""),
        "conversion": js.get("conversion", ""),
        "attempted": js.get("conversion_attempted"),
        "pillars": re.sub(r"\s+", " ", _section(text, "PILLARS CAPTURED")
                          or _grab(r"PILLARS CAPTURED:\s*([^\n]+)", text)).strip(),
        "happened": re.sub(r"\s+", " ", _section(text, "WHAT HAPPENED")).strip(),
        "fix": re.sub(r"\s+", " ", fix).strip(),
        "strengths": _numbered_items(_section(text, "TOP 3 STRENGTHS")),
        "improvements": _numbered_items(_section(text, "TOP 3 IMPROVEMENT AREAS")),
        "drill": re.sub(r"\s+", " ", drill).strip(),
    }


def parse_old_scores(old_dir: Path) -> dict[str, tuple[str, str]]:
    out: dict[str, tuple[str, str]] = {}
    for f in old_dir.glob("[0-9]*.md"):
        text = f.read_text(encoding="utf-8")
        tm = re.search(r"^\|\s*TOTAL[^|]*\|[^\n]*?([\d.]+)\s*/\s*100", text, re.M)
        band = _grab(r"Grade [Bb]and:\s*([^\n]+)", text).replace("*", "")
        out[re.match(r"^(\d+)", f.stem).group(1)] = (tm.group(1) if tm else "", band)
    return out


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 2)}"


def _est_height(text: str, width: int) -> float:
    lines = max(1, -(-len(text) // max(20, width - 5)))
    return min(110.0, 4 + lines * 13.5)


def _band_fill(cell) -> None:
    for band, color in BAND_FILLS.items():
        if cell.value and band.lower() in str(cell.value).lower():
            cell.fill = PatternFill("solid", fgColor=color)
            return


def _autowidth(ws, widths: dict[int, int]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def _num(v):
    if v is None:
        return ""
    try:
        return float(v)
    except (TypeError, ValueError):
        return v or ""


BODY = Font(size=10)
CENTER = Alignment(horizontal="center", vertical="center")


def _body_row(ws, ncols: int, center_cols=()) -> int:
    n = ws.max_row
    ws.row_dimensions[n].height = 15
    for c in range(1, ncols + 1):
        ws.cell(row=n, column=c).font = BODY
        if c in center_cols:
            ws.cell(row=n, column=c).alignment = CENTER
    return n


def _link_cell(ws, row: int, report: dict) -> None:
    cell = ws.cell(row=row, column=1)
    if report["link"]:
        cell.hyperlink = report["link"]
        cell.font = Font(color="0563C1", underline="single", size=10)


def _cat_score(r: dict, name: str, idx: int):
    v = r["cats_md"].get(name, "")
    if v in ("", None) and r["cats_js"]:
        v = r["cats_js"].get(str(idx))
        v = "N/A" if v is None and str(idx) in (r["cats_js"] or {}) else v
    return _num(v)


def build_workbook(reports: list[dict], old: dict[str, tuple[str, str]] | None) -> Workbook:
    wb = Workbook()
    full = [r for r in reports if r["call_type"] == "full"]
    short = [r for r in reports if r["call_type"] == "short"]

    cat_names: list[str] = []
    for r in full:
        for c in r["cats_md"]:
            if c not in cat_names:
                cat_names.append(c)
    cat_names.sort()

    ws = wb.active
    ws.title = "Summary"
    head = ["Call ID", "Caller", "Date/Time (UTC)", "Length", "Outcome", "Auto-Fail"] + \
        cat_names + ["Total /100", "Grade Band"]
    ws.append(head)
    for r in full:
        row = [r["call_id"], r["caller"], r["datetime"], r["length"],
               re.sub(r"\s+", " ", r["outcome"])[:80], r["autofail"]]
        row += [_cat_score(r, c, i + 1) for i, c in enumerate(cat_names)]
        row += [_num(r["total"]), r["band"]]
        ws.append(row)
        n = _body_row(ws, len(head), center_cols=tuple([4] + list(range(6, len(head) + 1))))
        for c in range(7, 7 + len(cat_names) + 1):
            ws.cell(row=n, column=c).number_format = "0.0"
        _link_cell(ws, n, r)
        _band_fill(ws.cell(row=n, column=len(head)))
    totals = [float(r["total"]) for r in full if r["total"] not in ("", None)]
    if totals:
        ws.append(["AVERAGE (full calls)"] + [""] * (len(head) - 3) +
                  [round(sum(totals) / len(totals), 1), ""])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10)
        c = ws.cell(row=ws.max_row, column=len(head) - 1)
        c.font = Font(bold=True, size=10)
        c.alignment = CENTER
    _style_header(ws, len(head))
    _autowidth(ws, {1: 11, 2: 16, 3: 18, 4: 7, 5: 42, 6: 9,
                    **{7 + i: 13 for i in range(len(cat_names))},
                    7 + len(cat_names): 9, 8 + len(cat_names): 12})

    if short:
        wsh = wb.create_sheet("Short Calls")
        FIX_W = 90
        head_s = ["Call ID", "Caller", "Date/Time (UTC)", "Length", "Outcome",
                  "Opener /5", "Conversion /5", "Attempt Made", "One Fix"]
        wsh.append(head_s)
        for r in short:
            att = r["attempted"]
            wsh.append([r["call_id"], r["caller"], r["datetime"], r["length"],
                        re.sub(r"\s+", " ", r["outcome"])[:60],
                        _num(r["opener"]), _num(r["conversion"]),
                        "" if att is None else ("YES" if att else "NO"),
                        r["fix"]])
            n = _body_row(wsh, len(head_s), center_cols=(4, 6, 7, 8))
            wsh.row_dimensions[n].height = max(15.0, _est_height(r["fix"], FIX_W))
            wsh.cell(row=n, column=9).alignment = Alignment(wrap_text=True, vertical="top")
            for c in (6, 7):
                wsh.cell(row=n, column=c).number_format = "0.0"
            _link_cell(wsh, n, r)
        op = [float(r["opener"]) for r in short if r["opener"] not in ("", None)]
        cv = [float(r["conversion"]) for r in short if r["conversion"] not in ("", None)]
        att_known = [r for r in short if r["attempted"] is not None]
        att_pct = (round(100 * sum(1 for r in att_known if r["attempted"]) / len(att_known))
                   if att_known else "")
        wsh.append(["AVERAGES", "", "", "", "",
                    round(sum(op) / len(op), 2) if op else "",
                    round(sum(cv) / len(cv), 2) if cv else "",
                    f"{att_pct}% attempted" if att_pct != "" else "", ""])
        for c in range(1, len(head_s) + 1):
            wsh.cell(row=wsh.max_row, column=c).font = Font(bold=True, size=10)
        _style_header(wsh, len(head_s))
        _autowidth(wsh, {1: 11, 2: 16, 3: 18, 4: 7, 5: 28, 6: 9, 7: 12, 8: 13, 9: FIX_W})

    wd = wb.create_sheet("Coaching Detail")
    DETAIL_W = 110
    head2 = ["Call ID", "Caller", "Type", "Total", "Band", "Section", "#", "Detail"]
    wd.append(head2)
    for r in reports:
        entries = []
        if r["pillars"]:
            entries.append(("Pillars", "", r["pillars"]))
        if r["happened"]:
            entries.append(("What happened", "", r["happened"]))
        entries += [("Strength", i, t) for i, t in enumerate(r["strengths"], 1)]
        entries += [("Improvement", i, t) for i, t in enumerate(r["improvements"], 1)]
        if r["fix"]:
            entries.append(("Fix", "", r["fix"]))
        if r["drill"]:
            entries.append(("Drill", "", r["drill"]))
        for sec, idx, txt in entries:
            wd.append([r["call_id"], r["caller"], r["call_type"],
                       _num(r["total"]), r["band"], sec, idx, txt])
            n = _body_row(wd, len(head2), center_cols=(3, 4, 7))
            wd.row_dimensions[n].height = _est_height(txt, DETAIL_W)
            dcell = wd.cell(row=n, column=8)
            dcell.alignment = Alignment(wrap_text=True, vertical="top")
            _band_fill(wd.cell(row=n, column=5))
    _style_header(wd, len(head2))
    _autowidth(wd, {1: 11, 2: 16, 3: 7, 4: 7, 5: 12, 6: 14, 7: 4, 8: DETAIL_W})

    wsc = wb.create_sheet("Caller Scorecard")
    head3 = ["Caller", "Full Calls", "Avg Total", "Min", "Max", "Recurring Lowest Category",
             "Short Calls", "Avg Opener /5", "Avg Conversion /5", "% Conversion Attempted"]
    wsc.append(head3)
    callers = sorted({r["caller"] for r in reports})
    for caller in callers:
        f = [r for r in full if r["caller"] == caller]
        s = [r for r in short if r["caller"] == caller]
        ts = [float(r["total"]) for r in f if r["total"] not in ("", None)]
        lows = []
        for r in f:
            scored = {c: _cat_score(r, c, i + 1) for i, c in enumerate(cat_names)}
            scored = {c: v for c, v in scored.items() if isinstance(v, float)}
            if scored:
                lows.append(min(scored, key=scored.get))
        op = [float(r["opener"]) for r in s if r["opener"] not in ("", None)]
        cv = [float(r["conversion"]) for r in s if r["conversion"] not in ("", None)]
        att_known = [r for r in s if r["attempted"] is not None]
        wsc.append([caller, len(f),
                    round(sum(ts) / len(ts), 1) if ts else "",
                    min(ts) if ts else "", max(ts) if ts else "",
                    Counter(lows).most_common(1)[0][0] if lows else "",
                    len(s),
                    round(sum(op) / len(op), 2) if op else "",
                    round(sum(cv) / len(cv), 2) if cv else "",
                    (f"{round(100 * sum(1 for r in att_known if r['attempted']) / len(att_known))}%"
                     if att_known else "")])
        _body_row(wsc, len(head3), center_cols=(2, 3, 4, 5, 7, 8, 9, 10))
    _style_header(wsc, len(head3))
    _autowidth(wsc, {1: 20, 2: 9, 3: 9, 4: 7, 5: 7, 6: 42, 7: 9, 8: 12, 9: 15, 10: 18})

    with_crits = [r for r in reports if r["crits"]]
    if len(with_crits) >= len(reports) / 2 and with_crits:
        crit_nums = sorted({k for r in with_crits for k in r["crits"]})
        names = {}
        for r in reports:
            names.update(r["crit_names"])
        wc = wb.create_sheet("Criterion Scores")
        wc.append(["Criterion", "Name"] + [r["call_id"] for r in reports])
        for num in crit_nums:
            wc.append([num, names.get(num, "")] +
                      [_num(r["crits"].get(num, "")) for r in reports])
            _body_row(wc, 2 + len(reports), center_cols=tuple(range(3, 3 + len(reports))))
        _style_header(wc, 2 + len(reports))
        _autowidth(wc, {1: 9, 2: 42, **{3 + i: 11 for i in range(len(reports))}})

    if old:
        wo = wb.create_sheet("Old vs New")
        wo.append(["Call ID", "Caller", "Old Total", "New Total", "Delta", "Old Band", "New Band"])
        for r in full:
            o_total, o_band = old.get(r["call_id"], ("", ""))
            delta = ""
            if o_total and r["total"] not in ("", None):
                delta = round(float(r["total"]) - float(o_total), 1)
            wo.append([r["call_id"], r["caller"], _num(o_total), _num(r["total"]),
                       delta, o_band, r["band"]])
            n = _body_row(wo, 7, center_cols=(3, 4, 5))
            _band_fill(wo.cell(row=n, column=6))
            _band_fill(wo.cell(row=n, column=7))
        _style_header(wo, 7)
        _autowidth(wo, {1: 11, 2: 16, 3: 10, 4: 10, 5: 8, 6: 13, 7: 13})

    return wb


def main() -> int:
    ap = argparse.ArgumentParser(description="Export call coaching reports to Excel")
    ap.add_argument("--dir", default=str(DEFAULT_DIR), help="reports directory to export")
    ap.add_argument("--out", help="output .xlsx path (default: <dir>/call_coaching_report.xlsx)")
    ap.add_argument("--compare", help="previous-run reports directory for an Old vs New sheet")
    ap.add_argument("--caller", help="only include reports whose filename contains this substring")
    args = ap.parse_args()

    rep_dir = Path(args.dir)
    files = sorted(f for f in rep_dir.glob("[0-9]*.md"))
    if args.caller:
        files = [f for f in files if args.caller.lower() in f.stem.lower()]
    if not files:
        print(f"No per-call reports found in {rep_dir}", file=sys.stderr)
        return 1
    reports = [parse_report(f) for f in files]
    old = parse_old_scores(Path(args.compare)) if args.compare else None

    out = Path(args.out) if args.out else rep_dir / "call_coaching_report.xlsx"
    build_workbook(reports, old).save(out)
    n_short = sum(1 for r in reports if r["call_type"] == "short")
    print(f"Exported {len(reports)} reports ({len(reports) - n_short} full, {n_short} short) -> {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
