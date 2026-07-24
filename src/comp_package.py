"""One-command comp package builder (the "158 Old State" document, generalized).

Pulls the subject + sold/active comps from the OpenWeb Ninja Zillow /search API
(via zillow_market_api), optionally clips them to a drawn boundary (bbox and/or
street regex), buckets comps by condition, builds a DUAL-TRACK ARV (same-bed
base case + higher-bed reconfig upside), runs the 4-tier rehab scenarios
(cosmetic / mid / full gut), computes wholesale MAO math off the CONSERVATIVE
track, matches local buyer-prospecting data by zip, and writes a branded Excel
workbook.

Dual-track ARV rule (Ty, 2026-07-21): a subject whose bedroom count is below
the comp set lives in a lower value band than the per-bedroom adjustment
implies. Base ARV comes ONLY from same-bed renovated comps; the higher-bed
track is a labeled upside that requires a verified reconfig, and underwriting
always uses the base track.

Usage:
  python src/comp_package.py --address "158 Old State Rd" --city Knoxville --zip 37914 \
      --beds 2 --baths 1 --sqft 1946 --year-built 1938 \
      --bbox "35.996,36.016,-83.895,-83.840" \
      --streets "old state|nash rd|seahorn|holston dr|bona|grata|silva|chilhowee"
"""

from __future__ import annotations

import argparse
import csv
import glob
import logging
import re
import statistics
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from comp_analyzer import fetch_subject_property
from rehab_estimator import estimate_rehab
from zillow_market_api import MarketListing, ZillowMarketAPI, filter_bbox, filter_streets

logger = logging.getLogger(__name__)

NAVY, BLUE, GREEN, GOLD = "0A1130", "316AFF", "1B9E5A", "B8860B"

GUT_ALLOWANCE_PER_SQFT = 15.0   # demo / drywall / insulation / reconfig framing
SOFT_COST_PCT = 0.13            # 3% permits + 10% contingency


# ── Condition bucketing ───────────────────────────────────────────────

def classify(listing: MarketListing) -> str:
    """Bucket a sold comp by condition using sold price vs Zestimate."""
    if listing.zestimate:
        ratio = listing.price / listing.zestimate
        if ratio >= 0.90:
            return "RENOVATED/RETAIL"
        if ratio <= 0.70:
            return "DISTRESSED"
        return "AVERAGE"
    return "UNKNOWN"


# ── Dual-track ARV ────────────────────────────────────────────────────

def dual_track_arv(subject_beds: int, subject_sqft: int,
                   sold: list[MarketListing]) -> dict:
    """Base ARV from same-bed renovated comps; upside from higher-bed band."""
    retail = [l for l in sold if classify(l) == "RENOVATED/RETAIL" and l.sqft and l.price]

    def track(comps: list[MarketListing], clamp_to_band: bool) -> dict:
        if not comps:
            return {}
        ppsf = statistics.median(l.ppsf for l in comps)
        est = ppsf * subject_sqft
        prices = sorted(l.price for l in comps)
        lo = prices[max(0, len(prices) // 4)]
        hi = prices[min(len(prices) - 1, (3 * len(prices)) // 4)]
        if clamp_to_band:
            # Bed-count bands cap regardless of size: an oversized subject
            # still sells to the same buyer pool, so never project past the
            # band's MEDIAN price. Analyst adjusts within the band for
            # micro-position (west vs east end of a pocket).
            est = min(est, statistics.median(prices))
        else:
            # Upside track: cap at the 75th percentile of the higher-bed band.
            est = min(est, hi)
        return {"arv": round(est / 5000) * 5000, "band": (lo, hi),
                "ppsf": round(ppsf), "n": len(comps)}

    same_bed = [l for l in retail if l.beds == subject_beds]
    base = track(same_bed, clamp_to_band=True)
    base_flag = ""
    if len(same_bed) < 3:
        wider = [l for l in retail if abs(l.beds - subject_beds) <= 1]
        base = track(wider, clamp_to_band=True)
        if base:
            base["arv"] = round(base["arv"] * 0.90 / 5000) * 5000
            base_flag = "THIN same-bed comp set - widened +/-1 bed and discounted 10%"

    upside_comps = [l for l in retail if l.beds > subject_beds
                    and (not subject_sqft or 0.6 <= (l.sqft / subject_sqft) <= 1.4)]
    upside = track(upside_comps, clamp_to_band=False)

    return {"base": base, "base_flag": base_flag, "upside": upside}


# ── Rehab scenarios ───────────────────────────────────────────────────

def rehab_scenarios(sqft: int, beds: int, baths: float, year_built: int) -> dict:
    """Cosmetic (wholetail) / mid (systems, no envelope) / full gut."""
    def total(est, drop=(), extra=0.0):
        rooms = [r for r in est.rooms if r.category not in drop]
        subtotal = sum(r.total for r in rooms) + extra
        return round(subtotal * (1 + SOFT_COST_PCT))

    cosmetic = estimate_rehab("", sqft, beds, baths, year_built,
                              tier=2, scope="wholetail", region="knoxville")
    full = estimate_rehab("", sqft, max(beds, 3), max(baths, 2.0), year_built,
                          tier=2, scope="full", region="knoxville")
    gut_extra = GUT_ALLOWANCE_PER_SQFT * sqft
    return {
        "cosmetic": total(cosmetic),
        "mid": total(full, drop=("Roof", "Windows", "Foundation/Structural")),
        "full_gut": total(full, extra=gut_extra),
    }


# ── Buyer matching ────────────────────────────────────────────────────

def local_buyers(zip_code: str) -> list[dict]:
    """Latest buyer-prospecting export rows tied to the subject zip."""
    paths = sorted(glob.glob(str(config.OUTPUT_DIR / "buyers_datasift_*.csv")))
    if not paths:
        return []
    rows = []
    with open(paths[-1], newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            blob = " ".join(str(v) for v in row.values())
            if zip_code and zip_code in blob:
                rows.append(row)
    return rows


# ── Workbook ──────────────────────────────────────────────────────────

def _header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)


def _title(ws, text, sub=""):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=14, color=NAVY)
    if sub:
        ws.cell(row=2, column=1, value=sub).font = Font(size=10, color="666666")


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(subject: dict, sold: list[MarketListing], active: list[MarketListing],
                   arv: dict, rehab: dict, buyers: list[dict], out_path: str) -> str:
    wb = Workbook()
    base, upside = arv.get("base") or {}, arv.get("upside") or {}
    base_arv = base.get("arv", 0)

    ws = wb.active
    ws.title = "Summary"
    _title(ws, f"{subject['address']} - Comp Package",
           f"Built {datetime.now():%m/%d/%Y}. Boundary-filtered API comps. "
           "Underwrite off the BASE (same-bed) track only.")
    mao = {k: {"mao70": round(base_arv * 0.70 - v), "mao75": round(base_arv * 0.75 - v)}
           for k, v in rehab.items()}
    rows = [
        ("SUBJECT", ""),
        ("Specs", f"{subject['sqft']:,} sqft | {subject['beds']}bd/{subject['baths']}ba | "
                  f"built {subject['year_built']}"),
        ("", ""),
        ("THE FOUR NUMBERS", ""),
        ("ARV - BASE (same-bed)", f"${base_arv:,.0f}" + (
            f" (band ${base.get('band', (0, 0))[0]:,.0f}-${base.get('band', (0, 0))[1]:,.0f}, "
            f"{base.get('n', 0)} comps @ ${base.get('ppsf', 0)}/sf median)" if base else " - NO COMPS")),
        ("ARV - UPSIDE (reconfig)", f"${upside.get('arv', 0):,.0f} "
         f"({upside.get('n', 0)} higher-bed comps) - only after a walkthrough verifies the "
         "layout converts" if upside else "n/a"),
        ("Full gut", f"${rehab['full_gut']:,.0f}"),
        ("Mid reno", f"${rehab['mid']:,.0f}"),
        ("Cosmetic", f"${rehab['cosmetic']:,.0f}"),
        ("", ""),
        ("WHOLESALE MATH (off BASE ARV)", ""),
        ("Full gut", f"MAO 70%: ${mao['full_gut']['mao70']:,.0f} | 75%: ${mao['full_gut']['mao75']:,.0f}"),
        ("Mid reno", f"MAO 70%: ${mao['mid']['mao70']:,.0f} | 75%: ${mao['mid']['mao75']:,.0f}"),
        ("Cosmetic", f"MAO 70%: ${mao['cosmetic']['mao70']:,.0f} | 75%: ${mao['cosmetic']['mao75']:,.0f}"),
    ]
    if arv.get("base_flag"):
        rows.append(("Flag", arv["base_flag"]))
    for i, r in enumerate(rows):
        ws.cell(row=4 + i, column=1, value=r[0])
        ws.cell(row=4 + i, column=2, value=r[1])
        if r[0] and not r[1]:
            ws.cell(row=4 + i, column=1).font = Font(bold=True, color=BLUE)
    _widths(ws, [28, 105])

    ws = wb.create_sheet("Sold Comps")
    _title(ws, "Sold comps (boundary-filtered)", "Bucketed by sold price vs Zestimate")
    hdr = ["Address", "Sold", "Price", "Bd", "Ba", "SqFt", "$/SF", "Bucket", "Zestimate", "URL"]
    for j, h in enumerate(hdr):
        ws.cell(row=4, column=j + 1, value=h)
    _header(ws, 4, len(hdr))
    order = {"RENOVATED/RETAIL": 0, "AVERAGE": 1, "UNKNOWN": 2, "DISTRESSED": 3}
    for i, l in enumerate(sorted(sold, key=lambda x: (order.get(classify(x), 9), x.sold_date), reverse=False)):
        bucket = classify(l)
        vals = [l.address, l.sold_date, l.price, l.beds, l.baths, l.sqft,
                round(l.ppsf) if l.ppsf else None, bucket, l.zestimate or None, l.detail_url]
        for j, v in enumerate(vals):
            ws.cell(row=5 + i, column=j + 1, value=v)
        ws.cell(row=5 + i, column=3).number_format = "$#,##0"
        ws.cell(row=5 + i, column=9).number_format = "$#,##0"
        color = {"RENOVATED/RETAIL": GREEN, "DISTRESSED": GOLD}.get(bucket)
        if color:
            ws.cell(row=5 + i, column=8).font = Font(bold=True, color=color, size=10)
    _widths(ws, [24, 11, 11, 5, 5, 7, 7, 18, 11, 45])

    ws = wb.create_sheet("Active")
    _title(ws, "Active listings (boundary-filtered)")
    hdr = ["Address", "List Price", "Bd", "Ba", "SqFt", "$/SF", "DOM", "URL"]
    for j, h in enumerate(hdr):
        ws.cell(row=4, column=j + 1, value=h)
    _header(ws, 4, len(hdr))
    for i, l in enumerate(active):
        vals = [l.address, l.price, l.beds, l.baths, l.sqft,
                round(l.ppsf) if l.ppsf else None, l.days_on_zillow, l.detail_url]
        for j, v in enumerate(vals):
            ws.cell(row=5 + i, column=j + 1, value=v)
        ws.cell(row=5 + i, column=2).number_format = "$#,##0"
    _widths(ws, [24, 12, 5, 5, 7, 7, 6, 45])

    ws = wb.create_sheet("Rehab Scenarios")
    _title(ws, f"Rehab scenarios - {subject['sqft']:,} sqft, built {subject['year_built']}",
           "SiftStack rehab engine, Knoxville multiplier, tier 2, incl. 13% soft costs. "
           f"Full gut adds ${GUT_ALLOWANCE_PER_SQFT:.0f}/sf demo-drywall allowance.")
    for i, (label, cost) in enumerate([("Cosmetic (wholetail, keeps bed count)", rehab["cosmetic"]),
                                       ("Mid reno (kitchen/baths/systems, no envelope)", rehab["mid"]),
                                       ("Full gut (everything + allowance)", rehab["full_gut"])]):
        ws.cell(row=4 + i, column=1, value=label)
        c = ws.cell(row=4 + i, column=2, value=cost)
        c.number_format = "$#,##0"
        ws.cell(row=4 + i, column=3, value=f"${cost / subject['sqft']:,.0f}/sf" if subject["sqft"] else "")
    _widths(ws, [46, 14, 10])

    if buyers:
        ws = wb.create_sheet("Buyer Targets")
        _title(ws, "Local buyer matches (latest buyer-prospecting export, zip-filtered)")
        hdr = ["Buyer", "Tags", "Notes"]
        for j, h in enumerate(hdr):
            ws.cell(row=4, column=j + 1, value=h)
        _header(ws, 4, len(hdr))
        for i, b in enumerate(buyers):
            ws.cell(row=5 + i, column=1, value=b.get("owner_name", ""))
            ws.cell(row=5 + i, column=2, value=b.get("tags", ""))
            ws.cell(row=5 + i, column=3, value=b.get("notes", ""))
        _widths(ws, [36, 30, 60])

    wb.save(out_path)
    return out_path


# ── CLI ───────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="Build a boundary-filtered comp package workbook")
    ap.add_argument("--address", required=True)
    ap.add_argument("--city", default="Knoxville")
    ap.add_argument("--state", default="TN")
    ap.add_argument("--zip", dest="zip_code", required=True)
    ap.add_argument("--beds", type=int, help="Override (county card beats Zillow)")
    ap.add_argument("--baths", type=float)
    ap.add_argument("--sqft", type=int)
    ap.add_argument("--year-built", type=int)
    ap.add_argument("--months", type=int, default=12, help="Sold lookback (default 12)")
    ap.add_argument("--bbox", help="lat_min,lat_max,lon_min,lon_max boundary box")
    ap.add_argument("--streets", help="Regex of in-boundary street names")
    ap.add_argument("--out", help="Output .xlsx path")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,
                        format="%(levelname)s %(message)s")

    subject_data = fetch_subject_property(args.address, args.city, args.state, args.zip_code)
    subject = {
        "address": f"{args.address}, {args.city}, {args.state} {args.zip_code}",
        "beds": args.beds or (subject_data.bedrooms if subject_data else 3),
        "baths": args.baths or (subject_data.bathrooms if subject_data else 2.0),
        "sqft": args.sqft or (subject_data.sqft if subject_data else 0),
        "year_built": args.year_built or (subject_data.year_built if subject_data else 0),
    }
    logger.info("Subject: %s | %s sqft %sbd/%sba built %s", subject["address"],
                subject["sqft"], subject["beds"], subject["baths"], subject["year_built"])

    api = ZillowMarketAPI()
    location = f"{args.city}, {args.state} {args.zip_code}"
    sold = api.pull_sold(location, months_back=args.months)
    active = api.pull_active(location)

    if args.bbox:
        lat_min, lat_max, lon_min, lon_max = (float(x) for x in args.bbox.split(","))
        sold = filter_bbox(sold, lat_min, lat_max, lon_min, lon_max)
        active = filter_bbox(active, lat_min, lat_max, lon_min, lon_max)
    if args.streets:
        pattern = re.compile(args.streets, re.I)
        sold = filter_streets(sold, pattern)
        active = filter_streets(active, pattern)
    logger.info("Boundary-filtered: %d sold, %d active", len(sold), len(active))
    if not sold:
        logger.error("No sold comps inside boundary - widen the boundary or months")
        return 1

    arv = dual_track_arv(subject["beds"], subject["sqft"], sold)
    rehab = rehab_scenarios(subject["sqft"], subject["beds"], subject["baths"],
                            subject["year_built"])
    buyers = local_buyers(args.zip_code)

    safe = "".join(c if c.isalnum() or c in " -" else "_" for c in args.address)[:40].strip().replace(" ", "_")
    out = args.out or str(Path.cwd() / f"{safe}_Comp_Package.xlsx")
    build_workbook(subject, sold, active, arv, rehab, buyers, out)

    base = arv.get("base") or {}
    print(f"\nARV base: ${base.get('arv', 0):,.0f} | upside: ${(arv.get('upside') or {}).get('arv', 0):,.0f}")
    print(f"Rehab: cosmetic ${rehab['cosmetic']:,.0f} | mid ${rehab['mid']:,.0f} | gut ${rehab['full_gut']:,.0f}")
    print(f"Workbook: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
