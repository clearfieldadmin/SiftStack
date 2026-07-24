"""Spec-driven deal-package workbook generator (the 158 Old State structure,
generalized). One clean Excel file per property with six sheets:

  1 Deal Summary    the numbers to use, value anchors, done-work story, gates
  2 Dial Sheet      ranked buyers with PER-BUYER open/target prices + contacts
  3 Deal Math       buyer-side + your-side math, rehab detail, dual-track ARV
  4 Comps           the sales that carry the story, each with its ROLE
  5 Pitch + Sequence 30-sec script, objection answers, day-by-day plan
  6 Sources + Audit  provenance for every number

Feed it a spec dict (see SPEC_TEMPLATE / build from a JSON file). Every section
is optional: missing keys are skipped, so a light deal renders a short book and
a full deal renders all six sheets. DataSift brand styling; zero em/en dashes.

Usage:
  python src/deal_package.py --spec deal_spec.json --out "123 Main St_Deal_Package.xlsx"
  python src/deal_package.py --template          # writes deal_spec_template.json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY, BLUE, GREEN, GOLD, RED = "0A1130", "316AFF", "1B9E5A", "B8860B", "CC0000"


# ── styling helpers ───────────────────────────────────────────────────

def _title(ws, text, sub=""):
    ws.cell(row=1, column=1, value=text).font = Font(bold=True, size=15, color=NAVY)
    if sub:
        ws.cell(row=2, column=1, value=sub).font = Font(size=10, color="666666")


def _header(ws, row, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _kv(ws, start, rows):
    """rows: list of [label, value]; a value of '' makes the label a blue subhead."""
    r = start
    for pair in rows:
        a, b = (pair + ["", ""])[:2]
        ws.cell(row=r, column=1, value=a)
        c = ws.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if a and not b:
            ws.cell(row=r, column=1).font = Font(bold=True, color=BLUE)
        r += 1
    return r


def _table(ws, start, rows, money_cols=()):
    r = start
    for row in rows:
        for j, v in enumerate(row, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if j in money_cols and isinstance(v, (int, float)):
                c.number_format = "$#,##0"
        r += 1
    return r


# ── sheet builders (each optional on its spec key) ────────────────────

def _sheet_summary(wb, spec):
    s = spec.get("summary")
    if not s:
        return
    ws = wb.create_sheet("1 Deal Summary")
    _title(ws, s.get("heading", f"{spec.get('address', 'Property')} - Deal Package"),
           s.get("subheading", ""))
    _kv(ws, 4, s.get("rows", []))
    _widths(ws, s.get("widths", [34, 120]))


def _sheet_dial(wb, spec):
    d = spec.get("dial_sheet")
    if not d:
        return
    ws = wb.create_sheet("2 Dial Sheet")
    _title(ws, d.get("heading", "Ranked buyer dial sheet with per-buyer pricing"),
           d.get("subheading", ""))
    cols = d.get("columns", ["#", "Buyer", "Person", "Best phone (score)", "More phones",
                             "Email", "Mail address", "Open $", "Target $", "Why them"])
    _header(ws, 4, cols)
    money = tuple(i + 1 for i, c in enumerate(cols) if "$" in c)
    _table(ws, 5, d.get("rows", []), money_cols=money)
    _widths(ws, d.get("widths", [4, 30, 34, 30, 26, 26, 32, 10, 10, 58]))


def _sheet_math(wb, spec):
    m = spec.get("deal_math")
    if not m:
        return
    ws = wb.create_sheet("3 Deal Math")
    _title(ws, m.get("heading", "Deal math + feasibility"), m.get("subheading", ""))
    _kv(ws, 4, m.get("rows", []))
    _widths(ws, m.get("widths", [40, 105]))


def _sheet_comps(wb, spec):
    c = spec.get("comps")
    if not c:
        return
    ws = wb.create_sheet("4 Comps")
    _title(ws, c.get("heading", "Comps that carry the story"), c.get("subheading", ""))
    cols = c.get("columns", ["Address", "Date", "Price", "Bd/Ba", "SqFt", "$/SF", "Role in the story"])
    _header(ws, 4, cols)
    money = tuple(i + 1 for i, col in enumerate(cols) if col.lower() in ("price", "$"))
    _table(ws, 5, c.get("rows", []), money_cols=money or (3,))
    _widths(ws, c.get("widths", [24, 11, 11, 7, 7, 7, 72]))


def _sheet_pitch(wb, spec):
    p = spec.get("pitch")
    if not p:
        return
    ws = wb.create_sheet("5 Pitch + Sequence")
    _title(ws, p.get("heading", "Scripts and the outreach plan"), p.get("subheading", ""))
    _kv(ws, 4, p.get("rows", []))
    _widths(ws, p.get("widths", [24, 125]))


def _sheet_sources(wb, spec):
    s = spec.get("sources")
    if not s:
        return
    ws = wb.create_sheet("6 Sources + Audit")
    _title(ws, s.get("heading", "Where every number came from"), s.get("subheading", ""))
    _kv(ws, 4, s.get("rows", []))
    _widths(ws, s.get("widths", [22, 128]))


BUILDERS = [_sheet_summary, _sheet_dial, _sheet_math, _sheet_comps, _sheet_pitch, _sheet_sources]


def build(spec: dict, out_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet; builders add their own
    for fn in BUILDERS:
        fn(wb, spec)
    if not wb.sheetnames:  # nothing in spec: leave a stub so save() works
        wb.create_sheet("Deal Package")
    wb.save(out_path)
    return out_path


SPEC_TEMPLATE = {
    "address": "123 Main St, City, ST 00000",
    "summary": {"subheading": "Final consolidated package.",
                "rows": [["THE NUMBERS TO USE", ""],
                         ["Contract with seller", "$X"],
                         ["Blast / asking price", "$Y"],
                         ["Target trade", "$..-$.."],
                         ["Floor", "$Z"]]},
    "dial_sheet": {"subheading": "3-source skip trace, Trestle scored.",
                   "rows": [[1, "Buyer LLC", "Person (age)", "(000) 000-0000 (100, x2)",
                             "backup", "email", "mail addr", 95000, 90000, "why them"]]},
    "deal_math": {"rows": [["BUYER-SIDE MATH", ""], ["At $Y", "profit / return"]]},
    "comps": {"rows": [["Anchor Rd", "2024-05-13", 120000, "3/1", 1749, 69, "the anchor"]]},
    "pitch": {"rows": [["30-SECOND PITCH", ""], ["Script", "..."]]},
    "sources": {"rows": [["Comps", "OpenWeb Ninja Zillow /search"],
                         ["Buyers", "SiftMap deed sweep"],
                         ["Skip trace", "Enformion + Tracerfy + web, Trestle scored"]]},
}


def main() -> int:
    ap = argparse.ArgumentParser(description="Generate a deal-package workbook from a spec")
    ap.add_argument("--spec", help="JSON spec file")
    ap.add_argument("--out", help="Output .xlsx path")
    ap.add_argument("--template", action="store_true", help="Write deal_spec_template.json and exit")
    args = ap.parse_args()

    if args.template:
        Path("deal_spec_template.json").write_text(json.dumps(SPEC_TEMPLATE, indent=1), encoding="utf-8")
        print("wrote deal_spec_template.json")
        return 0
    if not args.spec:
        ap.error("--spec is required (or use --template)")

    spec = json.load(open(args.spec, encoding="utf-8"))
    out = args.out or f"{Path(args.spec).stem}_Deal_Package.xlsx"
    build(spec, out)
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
